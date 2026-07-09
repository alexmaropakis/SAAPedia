"""File ingestion (CSV / TSV / XLSX): parse, map columns, de-duplicate, persist."""
from __future__ import annotations

import csv
import hashlib
import io
from dataclasses import dataclass, field

from sqlalchemy import delete as sa_delete
from sqlalchemy import or_, select
from sqlalchemy.orm import Session

from . import column_map
from .models import DatasetInfo, Observation, SAAP

# String value -> bool for flag columns.
_TRUE = {"yes", "true", "1", "y", "t"}
_FALSE = {"no", "false", "0", "n", "f"}


@dataclass
class IngestResult:
    filename: str
    rows_read: int = 0
    saap_created: int = 0
    observations_created: int = 0
    duplicate_observations_skipped: int = 0
    rows_skipped_no_identity: int = 0
    columns_mapped: dict[str, str] = field(default_factory=dict)
    columns_unmapped: list[str] = field(default_factory=list)
    dataset_dois_saved: int = 0
    saap_removed_no_uniprot: int = 0

    def as_dict(self) -> dict:
        return self.__dict__.copy()


def _to_float(value: str | None):
    if value is None:
        return None
    value = value.strip()
    if value == "" or value.lower() in {"na", "nan", "null", "none", "-"}:
        return None
    try:
        return float(value)
    except ValueError:
        return None


def _to_int(value: str | None):
    f = _to_float(value)
    return int(f) if f is not None else None


def _to_bool(value: str | None):
    if value is None:
        return None
    v = value.strip().lower()
    if v in _TRUE:
        return True
    if v in _FALSE:
        return False
    return None


def _clean_str(value: str | None):
    if value is None:
        return None
    v = value.strip()
    return v or None


def _saap_identity(record: dict) -> tuple[str, str, str] | None:
    """Identity key = (mtp_seq, bp_seq, aa_sub). Requires the substituted
    peptide at minimum; without it there is no SAAP to store."""
    mtp = _clean_str(record.get("mtp_seq"))
    if not mtp:
        return None
    return (mtp, _clean_str(record.get("bp_seq")) or "", _clean_str(record.get("aa_sub")) or "")


def _cell_to_str(v) -> str:
    """Normalize an XLSX cell value to the string form our parsers expect.
    Integer-valued floats (e.g. intensities read as 6.14e8) drop the '.0'."""
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v)


def _parse_delimited(raw_bytes: bytes, name: str) -> tuple[list[str], list[dict]]:
    text = raw_bytes.decode("utf-8-sig", errors="replace")
    delimiter = "\t" if name.endswith(".tsv") else ","
    reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
    headers = reader.fieldnames or []
    return headers, [dict(row) for row in reader]


def _parse_xlsx(raw_bytes: bytes) -> tuple[list[str], list[dict]]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
    try:
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            return [], []
        headers = [("" if h is None else str(h).strip()) for h in header_row]
        rows: list[dict] = []
        for r in rows_iter:
            if r is None or all(c is None for c in r):
                continue
            rec = {}
            for i, h in enumerate(headers):
                if not h:
                    continue
                rec[h] = _cell_to_str(r[i] if i < len(r) else None)
            rows.append(rec)
        return headers, rows
    finally:
        wb.close()


def _parse_file(raw_bytes: bytes, filename: str) -> tuple[list[str], list[dict]]:
    name = filename.lower()
    if name.endswith(".xlsx"):
        return _parse_xlsx(raw_bytes)
    return _parse_delimited(raw_bytes, name)


def ingest_file(
    db: Session,
    raw_bytes: bytes,
    filename: str,
    dataset_doi_map: dict[str, str] | None = None,
) -> IngestResult:
    headers, raw_rows = _parse_file(raw_bytes, filename)

    mapping, unmapped = column_map.map_headers(headers)
    result = IngestResult(
        filename=filename,
        columns_mapped=mapping,
        columns_unmapped=unmapped,
    )

    if dataset_doi_map:
        result.dataset_dois_saved = upsert_dataset_dois(db, dataset_doi_map)

    # Cache SAAP identities already seen in THIS ingest to avoid extra queries.
    saap_cache: dict[tuple[str, str, str], SAAP] = {}
    # Row hashes queued in THIS ingest (not yet flushed) so in-file duplicates
    # are collapsed instead of hitting the UNIQUE(row_hash) constraint at commit.
    seen_hashes: set[str] = set()

    for raw_row in raw_rows:
        result.rows_read += 1

        # Translate raw headers -> canonical fields.
        record: dict[str, str] = {}
        for raw_header, canonical in mapping.items():
            record[canonical] = raw_row.get(raw_header)

        identity = _saap_identity(record)
        if identity is None:
            result.rows_skipped_no_identity += 1
            continue

        saap = saap_cache.get(identity)
        if saap is None:
            saap = db.scalar(
                select(SAAP).where(
                    SAAP.mtp_seq == identity[0],
                    SAAP.bp_seq == identity[1],
                    SAAP.aa_sub == identity[2],
                )
            )
            if saap is None:
                saap = SAAP(
                    mtp_seq=identity[0],
                    bp_seq=identity[1],
                    aa_sub=identity[2],
                    source_accession=_clean_str(record.get("source_accession")),
                    source_gene=_clean_str(record.get("source_gene")),
                    ref_proteins=_clean_str(record.get("ref_proteins")),
                    immunoglobulin=_to_bool(record.get("immunoglobulin")),
                    trypsin=_to_bool(record.get("trypsin")),
                    missed_cleavage=_to_bool(record.get("missed_cleavage")),
                    aas_at_peptide_terminus=_to_bool(record.get("aas_at_peptide_terminus")),
                    greater_than_shared=_to_bool(record.get("greater_than_shared")),
                )
                db.add(saap)
                db.flush()  # assign PK
                result.saap_created += 1
            else:
                # Backfill source metadata if the existing row lacks it.
                _backfill_source(saap, record)
            saap_cache[identity] = saap
        else:
            _backfill_source(saap, record)

        row_hash = _hash_row(identity, record)
        if row_hash in seen_hashes or db.scalar(
            select(Observation.id).where(Observation.row_hash == row_hash)
        ) is not None:
            result.duplicate_observations_skipped += 1
            continue
        seen_hashes.add(row_hash)

        db.add(Observation(
            saap_id=saap.id,
            dataset=_clean_str(record.get("dataset")),
            tmt_tissue=_clean_str(record.get("tmt_tissue")),
            digest=_clean_str(record.get("digest")),
            species=_clean_str(record.get("species")),
            acquisition_type=_clean_str(record.get("acquisition_type")),
            saap_pep=_to_float(record.get("saap_pep")),
            positional_probability=_to_float(record.get("positional_probability")),
            n_evidence_fragments=_to_int(record.get("n_evidence_fragments")),
            source_file=filename,
            row_hash=row_hash,
        ))
        result.observations_created += 1

    db.commit()

    # A peptide must have a UniProt accession; drop those that don't.
    result.saap_removed_no_uniprot = cleanup_saap_without_uniprot(db)
    return result


def cleanup_saap_without_uniprot(db: Session) -> int:
    """Delete SAAP that have no UniProt accession. Returns the count removed."""
    victims = db.scalars(
        select(SAAP.id).where(or_(SAAP.source_accession.is_(None),
                                  SAAP.source_accession == ""))
    ).all()
    if not victims:
        return 0
    victim_list = list(victims)
    db.execute(sa_delete(Observation).where(Observation.saap_id.in_(victim_list)))
    db.execute(sa_delete(SAAP).where(SAAP.id.in_(victim_list)))
    db.commit()
    return len(victim_list)


def upsert_dataset_dois(db: Session, dataset_doi_map: dict[str, str]) -> int:
    """Insert or update dataset -> DOI rows. Blank names are ignored; a blank
    DOI clears any existing one. Returns the number of datasets written."""
    saved = 0
    for name, doi in dataset_doi_map.items():
        name = (name or "").strip()
        if not name:
            continue
        doi = (doi or "").strip() or None
        existing = db.scalar(select(DatasetInfo).where(DatasetInfo.name == name))
        if existing is None:
            db.add(DatasetInfo(name=name, doi=doi))
        else:
            existing.doi = doi
        saved += 1
    db.commit()
    return saved


def _backfill_source(saap: SAAP, record: dict) -> None:
    """Fill in per-peptide attributes on an existing SAAP if previously blank."""
    for attr in ("source_accession", "source_gene", "ref_proteins"):
        if getattr(saap, attr) is None:
            val = _clean_str(record.get(attr))
            if val:
                setattr(saap, attr, val)
    for attr in ("immunoglobulin", "trypsin", "missed_cleavage",
                 "aas_at_peptide_terminus", "greater_than_shared"):
        if getattr(saap, attr) is None:
            b = _to_bool(record.get(attr))
            if b is not None:
                setattr(saap, attr, b)


def _hash_row(identity: tuple[str, str, str], record: dict) -> str:
    """Stable hash over identity + all observation fields, so re-uploading the
    exact same line is recognized as a duplicate observation."""
    parts = list(identity)
    for f in sorted(column_map.OBSERVATION_FIELDS):
        parts.append(f"{f}={(record.get(f) or '').strip()}")
    return hashlib.sha1("\x1f".join(parts).encode("utf-8")).hexdigest()
